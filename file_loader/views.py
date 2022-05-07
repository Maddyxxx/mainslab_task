from django.shortcuts import render
from django.db.utils import IntegrityError
import openpyxl
from rest_framework import views
from clients.models import Client, Organization
from bills.models import Bill


def get_table_data(excel_file):
    data = {}
    sheet_names = []
    wb = openpyxl.load_workbook(excel_file)
    sheets = wb.sheetnames
    for sheet in sheets:
        sheet_names.append(sheet)
        worksheet = wb[sheet]
        data[sheet] = []
        for row in worksheet.iter_rows():
            row_data = [i.value for i in row if i.value is not None]
            data[sheet].append(row_data)
        data[sheet] = data[sheet][1:]
    return data, sheet_names


class LoadClientOrg(views.View):

    def get(self, request):
        return render(request, 'load_client_org.html', {})

    def post(self, request):
        excel_file = request.FILES["excel_file"]
        table_data, sheet_names = get_table_data(excel_file)
        if len(sheet_names) < 2:
            return render(request, 'load_bills.html', {'text': 'Ошибка при загрузке файла'})

        clients = table_data[sheet_names[0]]
        for client in clients:
            new_client = Client(client_name=client[0])
            try:
                new_client.save()
            except IntegrityError:
                pass
        organizations = table_data[sheet_names[1]]
        for organization in organizations:
            client = Client.objects.get(client_name=organization[0])
            if client:
                org = Organization(
                    client=client,
                    organization_name=organization[1]
                )
                try:
                    org.save()
                except IntegrityError:
                    pass
        return render(request, 'load_client_org.html', {'text': 'Файл успешно загружен'})


class LoadBills(views.View):

    def get(self, request):
        return render(request, 'load_bills.html', {})

    def post(self, request):
        excel_file = request.FILES["excel_file"]
        table_data, sheet_names = get_table_data(excel_file)
        bills = table_data[sheet_names[0]]
        for bill in bills:
            organization_name = bill[0]
            client_org = Organization.objects.filter(organization_name=organization_name)

            if client_org:
                bill_number, bill_sum, bill_date = bill[1], bill[2], bill[3]
                new_bill = Bill(
                    client_org=client_org[0],
                    bill_number=bill_number,
                    bill_sum=bill_sum,
                    bill_date=bill_date
                )
                try:
                    new_bill.save()
                except IntegrityError:
                    pass
            else:
                return render(request, 'load_bills.html', {'text': 'Ошибка при загрузке файла'})

        return render(request, 'load_bills.html', {'text': 'Файл успешно загружен'})
